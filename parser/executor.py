from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pdfplumber

from schemas import DOC_TYPE_MAP

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".tiff"}
NUMERIC_FIELDS = {"amount", "debit", "credit", "balance", "revenue", "net_income"}


def execute_parser(file_path: Path, dsl: dict, doc_type: str) -> dict:
    ext = file_path.suffix.lower()

    if ext in IMAGE_EXTS:
        data = _parse_image(file_path, dsl)
    elif ext == ".pdf":
        data = _parse_pdf(file_path, dsl)
    elif ext == ".txt":
        data = _parse_text(file_path, dsl)
    elif ext in {".xlsx", ".xls"}:
        data = _parse_excel(file_path, dsl)
    else:
        data = _apply_regex_fields("", dsl.get("fields", {}))

    model = DOC_TYPE_MAP[doc_type]()
    validated = model.model_validate({**model.model_dump(), **data})
    return validated.model_dump()


def _coerce_value(field: str, val) -> str | float:
    if field in NUMERIC_FIELDS:
        try:
            return float(str(val).replace("$", "").replace(",", "").strip())
        except (ValueError, TypeError):
            return 0.0
    return str(val).strip() if val is not None else ""


def _parse_image(file_path: Path, dsl: dict) -> dict:
    extracted = dsl.get("extracted")
    if isinstance(extracted, dict) and _has_values(extracted):
        return {
            k: _coerce_value(k, v)
            for k, v in extracted.items()
            if not isinstance(v, dict)
        }

    ocr_text = _ocr_text(file_path)
    if ocr_text:
        return _apply_regex_fields(ocr_text, dsl.get("fields", {}))

    return {}


def _has_values(data: dict) -> bool:
    return any(v not in (None, "", {}) and not isinstance(v, dict) for v in data.values())


def _ocr_text(file_path: Path) -> str:
    try:
        import pytesseract
        from PIL import Image

        return pytesseract.image_to_string(Image.open(file_path))
    except Exception:
        return ""


def _apply_regex_fields(text: str, fields: dict) -> dict:
    result = {}
    for field, spec in fields.items():
        if isinstance(spec, dict) and spec.get("method") == "regex":
            try:
                match = re.search(spec["pattern"], text, re.IGNORECASE | re.MULTILINE)
                if match:
                    raw = match.group(1).strip() if match.lastindex else match.group(0).strip()
                else:
                    raw = ""
            except (IndexError, re.error):
                raw = ""
            result[field] = _coerce_value(field, raw)
        elif isinstance(spec, str):
            result[field] = _coerce_value(field, spec)
    return result


def _merge_extracted(result: dict, dsl: dict) -> dict:
    extracted = dsl.get("extracted")
    if not isinstance(extracted, dict):
        return result
    merged = dict(result)
    for field, val in extracted.items():
        if isinstance(val, dict):
            continue
        if field not in merged or merged[field] in ("", 0, 0.0):
            merged[field] = _coerce_value(field, val)
    return merged


def _parse_text(file_path: Path, dsl: dict) -> dict:
    text = file_path.read_text(encoding="utf-8", errors="replace")
    result = _apply_regex_fields(text, dsl.get("fields", {}))
    return _merge_extracted(result, dsl)


def _parse_pdf(file_path: Path, dsl: dict) -> dict:
    text = ""
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages[:3]:
            text += page.extract_text() or ""
    result = _apply_regex_fields(text, dsl.get("fields", {}))
    return _merge_extracted(result, dsl)


def _parse_excel(file_path: Path, dsl: dict) -> dict:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_index = dsl.get("sheet", 0)
    sheet = wb.worksheets[sheet_index]
    columns = dsl.get("columns", {})
    result = {}

    header_row = 1
    col_map = {}
    for col_idx in range(1, sheet.max_column + 1):
        val = sheet.cell(header_row, col_idx).value
        if val is not None:
            col_map[str(val).strip().lower()] = col_idx

    data_row = 2
    for field, col_spec in columns.items():
        if isinstance(col_spec, str) and len(col_spec) == 1 and col_spec.isalpha():
            col_idx = openpyxl.utils.column_index_from_string(col_spec)
            val = sheet.cell(data_row, col_idx).value
        elif isinstance(col_spec, str):
            col_idx = col_map.get(col_spec.lower(), 0)
            val = sheet.cell(data_row, col_idx).value if col_idx else ""
        else:
            val = ""
        result[field] = _coerce_value(field, val if val is not None else "")

    return result
